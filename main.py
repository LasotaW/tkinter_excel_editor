import tkinter as tk
from tkinter import StringVar, filedialog, ttk
from openpyxl import load_workbook

class App(tk.Frame):
    def __init__(self, master=None):
        tk.Frame.__init__(self, master)
        self.master = master
        self.pack()

        self.exitButton = ttk.Button(root, text="Wybierz plik", command=self.browseFile, width=25)
        self.exitButton.pack()

    def browseFile(self):
        self.file = filedialog.askopenfile(parent=root, title="Wybierz plik", filetypes =[('Excel Files', '*.xlsx *.xlsm *.sxc *.ods *.csv *.tsv')])
        self.wb = load_workbook(filename=self.file.name)
        self.chosenFile = ttk.Label(root, text="Wybrany plik: " + self.file.name)
        self.chosenFile.pack()

        self.wb = load_workbook(self.file.name)
        self.availableSheets = ttk.Label(root, text="Dostępne arkusze: ")
        self.availableSheets.pack(pady=5)

        self.option = StringVar()
        self.sheetCombo = ttk.Combobox(root, textvariable=self.option, values=self.wb.sheetnames, width=22)
        self.sheetCombo.current(0)
        self.sheetCombo.pack()

        self.changeButton = ttk.Button(root, text="Zamień dane/wypełnij", command=self.manipulateData, width=25)
        self.changeButton.pack(pady=5)

        self.mergeButton = ttk.Button(root, text="Połącz z drugim plikiem", command=self.mergeWindow, width=25)
        self.mergeButton.pack()

    def manipulateData(self):
        self.newWindow = tk.Toplevel(root)
        self.newWindow.geometry("400x300+50+50")
        self.newWindow.resizable(False, False)

        self.ws = self.wb[self.sheetCombo.get()]

        self.lbl = ttk.Label(self.newWindow, text="Wartość do wprowadzenia: ")
        self.usrInput = ttk.Entry(self.newWindow, width=50)
        self.lbl.pack(pady=5)
        self.usrInput.pack()

        self.cordLabel = ttk.Label(self.newWindow, text="Współrzędne:")
        self.cordLabel.pack(pady=10)
        self.cordEntry = ttk.Entry(self.newWindow)
        self.cordEntry.pack()   

        self.applyButton = ttk.Button(self.newWindow, text="Potwierdź", command=self.apply, width=20)
        self.applyButton.pack(pady=5)

    def apply(self):
        coord = self.cordEntry.get()
        range = self.ws[coord]
        try:
            try:
                for i in range:
                    for j in i:
                        j.value = self.usrInput.get()
            except:
                for i in range:
                    i.value = self.usrInput.get()
        except:
            self.ws[coord] = self.usrInput.get()

        try:
            self.wb.save(self.file.name)
            self.successLabel = tk.Label(self.newWindow, text="Pomyślnie zapisano dane!", fg='green')
            self.successLabel.pack()

        except:
            print("Błąd!")

    def mergeWindow(self):
        self.newWindow = tk.Toplevel(root)
        self.newWindow.geometry("400x150+50+50")
        self.newWindow.resizable(False, False)

        self.fileTwo = filedialog.askopenfile(parent=self.newWindow, title="Wybierz plik", filetypes =[('Excel Files', '*.xlsx *.xlsm *.sxc *.ods *.csv *.tsv')])
        self.wbTwo = load_workbook(filename=self.fileTwo.name)
        self.chosenFileTwo = ttk.Label(self.newWindow, text="Wybrany plik: " + self.fileTwo.name)
        self.chosenFileTwo.pack()

        self.wbTwo = load_workbook(self.fileTwo.name)
        self.availableSheetsTwo = ttk.Label(self.newWindow, text="Dostępne arkusze: ")
        self.availableSheetsTwo.pack(pady=10)

        self.optionTwo = StringVar()
        self.sheetComboTwo = ttk.Combobox(self.newWindow, textvariable=self.optionTwo, values=self.wbTwo.sheetnames)
        self.sheetComboTwo.current(0)
        self.sheetComboTwo.pack()

        self.actionButtonTwo = ttk.Button(self.newWindow, text="Połącz dane", command=self.merge, width=22)
        self.actionButtonTwo.pack(pady=10)

    def merge(self):
        self.wsTwo = self.wbTwo[self.sheetComboTwo.get()]
        self.ws = self.wb[self.sheetCombo.get()]
        offset = self.ws.max_column

        for i in range(1, self.wsTwo.max_column):
            for j in range(1, self.wsTwo.max_row):
                self.ws.cell(column=offset+i, row=j, value=self.wsTwo.cell(column=i, row=j).value)

        try:
            self.wb.save(self.file.name)
            self.successLabel = tk.Label(self.newWindow, text="Pomyślnie dołączono dane!", fg='green')
            self.successLabel.pack()
        except:
            print("ERROR")
        

root = tk.Tk()
myapp = App(root)

myapp.master.title("Excel multitool")
root.geometry("400x300+50+50")
root.resizable(False,False)
root.columnconfigure(0, weight=1)

myapp.mainloop()